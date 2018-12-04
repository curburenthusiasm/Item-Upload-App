
# coding: utf-8

# In[1]:


# import openpyxl and tkinter modules 
from openpyxl import *
from Tkinter import *
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('C:\Users\Rfoley\MNI Import.xlsx') 
  
# create the sheet object 
sheet = wb.active 
  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30
  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "ItemCode"
    sheet.cell(row=1, column=2).value = "ItemCodeDesc"
    sheet.cell(row=1, column=3).value = "ProductLine"
    sheet.cell(row=1, column=4).value = "StandardUnitCost"
    sheet.cell(row=1, column=5).value = "StandardPrice"
    sheet.cell(row=1, column=6).value = "ShipWeight"
    sheet.cell(row=1, column=7).value = "UDF_Dimensions"
    sheet.cell(row=1, column=8).value = "UDF_SIZE"
    sheet.cell(row=1, column=9).value = "UDF_CROSSREF"
  
  
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the ItemCode box 
    ItemCode.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the ItemCodeDesc box 
    ItemCodeDesc.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the ProductLine box 
    ProductLine.focus_set() 
  
  
# Function to set focus 
def focus4(event): 
    # set focus on the StandardUnitCost box 
    StandardUnitCost.focus_set() 
  
  
# Function to set focus 
def focus5(event): 
    # set focus on the StandardPrice box 
    StandardPrice.focus_set() 
  
  
# Function to set focus 
def focus6(event): 
    # set focus on the ShipWeight box 
    ShipWeight.focus_set() 

# Function to set focus 
def focus7(event): 
    # set focus on the UDF_DIMENSIONS box 
    UDF_DIMENSIONS.focus_set()

# Function to set focus 
def focus8(event): 
    # set focus on the UDF_SIZE box 
    UDF_SIZE.focus_set()
    
# Function to set focus 
def focus9(event): 
    # set focus on the UDF_CROSSREF box 
    UDF_CROSSREF.focus_set()
  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    ItemCode.delete(0, END) 
    ItemCodeDesc.delete(0, END) 
    ProductLine.delete(0, END) 
    StandardUnitCost.delete(0, END) 
    StandardPrice.delete(0, END) 
    Shipweight.delete(0, END) 
    UDF_DIMENSIONS.delete(0, END)
    UDF_SIZE.delete(0, END)
    UDF_CROSSREF.delete(0, END)
    
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (ItemCode.get() == "" and
        ItemCodeDesc.get() == "" and
        ProductLine.get() == "" and
        StandardUnitCost.get() == "" and
        StandardPrice.get() == "" and
        Shipweight.get() == "" and
        UDF_DIMENSIONS.get() == "" and
        UDF_SIZE.get() == "" and
        UDF_CROSSREF.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = ItemCode.get() 
        sheet.cell(row=current_row + 1, column=2).value = ItemCodeDesc.get() 
        sheet.cell(row=current_row + 1, column=3).value = ProductLine.get() 
        sheet.cell(row=current_row + 1, column=4).value = StandardUnitCost.get() 
        sheet.cell(row=current_row + 1, column=5).value = StandardPrice.get() 
        sheet.cell(row=current_row + 1, column=6).value = Shipweight.get() 
        sheet.cell(row=current_row + 1, column=7).value = UDF_DIMENSIONS.get() 
        sheet.cell(row=current_row + 1, column=8).value = UDF_SIZE.get()
        sheet.cell(row=current_row + 1, column=9).value = UDF_CROSSREF.get()
        
        # save the file 
        wb.save('C:\Users\Rfoley\MNI Import.xlsx') 
  
        # set focus on the name_field box 
        ItemCode.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("New Master Number Upload") 
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
    excel() 
  
    # create a ItemCode label 
    ItemCode = Label(root, text="ItemCode", bg="light green") 
  
    # create a ItemCodeDesc label 
    ItemCodeDesc = Label(root, text="ItemCodeDesc", bg="light green") 
  
    # create a ProductLine label 
    ProductLine = Label(root, text="ProductLine", bg="light green") 
  
    # create a StandardUnitCost label 
    StandardUnitCost = Label(root, text="StandardUnitCost", bg="light green") 
  
    # create a StandardPrice lable 
    StandardPrice = Label(root, text="StandardPrice", bg="light green") 
  
    # create a ShipWeight label 
    Shipweight = Label(root, text="Shipweight.", bg="light green") 
  
    # create a UDF_DIMENSIONS label 
    UDF_DIMENSIONS = Label(root, text="UDF_DIMENSIONS", bg="light green") 
  
    # create a UDF_SIZE label 
    UDF_SIZE = Label(root, text="UDF_SIZE", bg="light green")
    
    # create a UDF_CROSSREF label 
    UDF_CROSSREF = Label(root, text="UDF_CROSSREF", bg="light green")
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    
    ItemCode.grid(row=1, column=0) 
    ItemCodeDesc.grid(row=2, column=0) 
    ProductLine.grid(row=3, column=0) 
    StandardUnitCost.grid(row=4, column=0) 
    StandardPrice.grid(row=5, column=0) 
    Shipweight.grid(row=6, column=0) 
    UDF_DIMENSIONS.grid(row=7, column=0)
    UDF_SIZE.grid(row=8, column=0)
    UDF_CROSSREF.grid(row=9, column=0)
  
    # create a text entry box 
    # for typing the information 
    ItemCode = Entry(root) 
    ItemCodeDesc = Entry(root) 
    ProductLine = Entry(root) 
    StandardUnitCost = Entry(root) 
    StandardPrice = Entry(root)
    Shipweight = Entry(root)
    UDF_DIMENSIONS = Entry(root) 
    UDF_SIZE = Entry(root)
    UDF_CROSSREF = Entry(root)
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    ItemCode.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    ItemCodeDesc.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    ProductLine.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    StandardUnitCost.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    StandardPrice.bind("<Return>", focus5) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    Shipweight.bind("<Return>", focus6)
    
     # whenever the enter key is pressed 
    # then call the focus6 function 
    UDF_DIMENSIONS.bind("<Return>", focus7)
    
     # whenever the enter key is pressed 
    # then call the focus6 function 
    UDF_SIZE.bind("<Return>", focus8)
    
         # whenever the enter key is pressed 
    # then call the focus6 function 
    UDF_CROSSREF.bind("<Return>", focus9)
    
    
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    ItemCode.grid(row=1, column=1, ipadx="100") 
    ItemCodeDesc.grid(row=2, column=1, ipadx="100") 
    ProductLine.grid(row=3, column=1, ipadx="100") 
    StandardUnitCost.grid(row=4, column=1, ipadx="100") 
    StandardPrice.grid(row=5, column=1, ipadx="100") 
    Shipweight.grid(row=6, column=1, ipadx="100") 
    UDF_DIMENSIONS.grid(row=7, column=1, ipadx="100")
    UDF_SIZE.grid(row=8, column=1, ipadx="100")
    UDF_CROSSREF.grid(row=9, column=1, ipadx="100")
  
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=10, column=1) 
  
    # start the GUI 
    root.mainloop()

