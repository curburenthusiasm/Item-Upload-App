
# coding: utf-8

# In[12]:


# import openpyxl and tkinter modules 
from openpyxl import *
from Tkinter import *


# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('C:\Users\Public\Alias Item Upload.xlsx') 
  
# create the sheet object 
sheet = wb.active 
  
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30

  
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "ItemCode"
    sheet.cell(row=1, column=2).value = "AliasItemCode"
    sheet.cell(row=1, column=3).value = "AliasDescription"
    sheet.cell(row=1, column=4).value = "Sage100CustNumber"
    sheet.cell(row=1, column=5).value = "SpecialPrice"
    sheet.cell(row=1, column=6).value = "UPC"
    sheet.cell(row=1, column=7).value = "Brand"
  
  
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the ItemCode box 
    ItemCode.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the AliasItemCode box 
    AliasItemCode.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the AliasDescription box 
    AliasDescription.focus_set() 
  
  
# Function to set focus 
def focus4(event): 
    # set focus on the Sage100CustNumber box 
    Sage100CustNumber.focus_set() 
  
  
# Function to set focus 
def focus5(event): 
    # set focus on the SpecialPrice box 
    SpecialPrice.focus_set() 
  
  
# Function to set focus 
def focus6(event): 
    # set focus on the UPC box 
    UPC.focus_set() 

# Function to set focus 
def focus7(event): 
    # set focus on the Brand box 
    Brand.focus_set()


  
  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    ItemCode.delete(0, END) 
    AliasItemCode.delete(0, END) 
    AliasDescription.delete(0, END) 
    Sage100CustNumber.delete(0, END) 
    SpecialPrice.delete(0, END) 
    UPC.delete(0, END) 
    Brand.delete(0, END)
    
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (ItemCode.get() == "" and
        AliasItemCode.get() == "" and
        AliasDescription.get() == "" and
        Sage100CustNumber.get() == "" and
        SpecialPrice.get() == "" and
        UPC.get() == "" and
        Brand.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = ItemCode.get() 
        sheet.cell(row=current_row + 1, column=2).value = AliasItemCode.get() 
        sheet.cell(row=current_row + 1, column=3).value = AliasDescription.get() 
        sheet.cell(row=current_row + 1, column=4).value = Sage100CustNumber.get() 
        sheet.cell(row=current_row + 1, column=5).value = SpecialPrice.get() 
        sheet.cell(row=current_row + 1, column=6).value = UPC.get() 
        sheet.cell(row=current_row + 1, column=7).value = Brand.get()
        
        
        # save the file 
        wb.save('C:\Users\Public\Alias Item Upload.xlsx') 
  
        # set focus on the name_field box 
        ItemCode.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light blue') 
  
    # set the title of GUI window 
    root.title("Alias Item Upload") 
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
    excel() 
  
    # create a ItemCode label 
    ItemCode = Label(root, text="ItemCode", bg="light blue") 
  
    # create a AliasItemCode label 
    AliasItemCode = Label(root, text="AliasItemCode", bg="light blue") 
  
    # create a AliasDescription label 
    AliasDescription = Label(root, text="AliasDescription", bg="light blue") 
  
    # create a Sage100CustNumber label 
    Sage100CustNumber = Label(root, text="Sage100CustNumber", bg="light blue") 
  
    # create a SpecialPrice lable 
    SpecialPrice = Label(root, text="SpecialPrice", bg="light blue") 
  
    # create a UPC label 
    UPC = Label(root, text="UPC.", bg="light blue") 
  
    # create a Brand label 
    Brand = Label(root, text="Brand", bg="light blue") 
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    
    ItemCode.grid(row=1, column=0) 
    AliasItemCode.grid(row=2, column=0) 
    AliasDescription.grid(row=3, column=0) 
    Sage100CustNumber.grid(row=4, column=0) 
    SpecialPrice.grid(row=5, column=0) 
    UPC.grid(row=6, column=0) 
    Brand.grid(row=7, column=0)

    # create a text entry box 
    # for typing the information 
    ItemCode = Entry(root) 
    AliasItemCode = Entry(root) 
    AliasDescription = Entry(root) 
    Sage100CustNumber = Entry(root) 
    SpecialPrice = Entry(root)
    UPC = Entry(root)
    Brand = Entry(root) 
  
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    ItemCode.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    AliasItemCode.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    AliasDescription.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    Sage100CustNumber.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    SpecialPrice.bind("<Return>", focus5) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    UPC.bind("<Return>", focus6)
    
     # whenever the enter key is pressed 
    # then call the focus6 function 
    Brand.bind("<Return>", focus7)    
    
  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    ItemCode.grid(row=1, column=1, ipadx="100") 
    AliasItemCode.grid(row=2, column=1, ipadx="100") 
    AliasDescription.grid(row=3, column=1, ipadx="100") 
    Sage100CustNumber.grid(row=4, column=1, ipadx="100") 
    SpecialPrice.grid(row=5, column=1, ipadx="100") 
    UPC.grid(row=6, column=1, ipadx="100") 
    Brand.grid(row=7, column=1, ipadx="100")
  
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=10, column=1) 
  

    # start the GUI 
    root.mainloop()

